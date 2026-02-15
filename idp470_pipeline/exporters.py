from __future__ import annotations

from datetime import datetime
from io import BytesIO
import logging
import re
from collections import Counter
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd

from .models import ContractSpec

LOGGER = logging.getLogger(__name__)

_DEFAULT_RECORD_ORDER = ["FIC", "ENT", "ECH", "COM", "REF", "ADR", "AD2", "LIG", "LEC", "PIE"]
_SHEET_SUMMARY = "SYNTHESE"
_SHEET_ALL = "TOUS"
_SHEET_DICTIONARY = "DICTIONNAIRE"


def _record_order_map(contract: ContractSpec | None) -> dict[str, int]:
    order_map: dict[str, int] = {}
    if contract is not None and contract.structure_rules:
        for rule in sorted(contract.structure_rules, key=lambda item: item.order_index):
            record_name = str(rule.record_name).strip().upper()
            if record_name and record_name not in order_map:
                order_map[record_name] = int(rule.order_index)

    if not order_map:
        for idx, name in enumerate(_DEFAULT_RECORD_ORDER, start=1):
            order_map[name] = idx
    return order_map


def _sort_record_types(record_types: list[str], contract: ContractSpec | None) -> list[str]:
    order_map = _record_order_map(contract)
    normalized = [str(record_type).strip().upper() for record_type in record_types if str(record_type).strip()]
    return sorted(
        normalized,
        key=lambda record_type: (order_map.get(record_type, 10_000), record_type),
    )


def _safe_table_name(base: str, index: int) -> str:
    cleaned = re.sub(r"[^0-9a-zA-Z_]", "_", base)
    if not cleaned or not cleaned[0].isalpha():
        cleaned = f"T_{cleaned}"
    return f"{cleaned}_{index}"


def _safe_sheet_name(base: str, existing: set[str]) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", str(base).strip())
    if not cleaned:
        cleaned = "TYPE"
    candidate = cleaned[:31]
    index = 1
    while candidate in existing:
        suffix = f"_{index}"
        candidate = f"{cleaned[: max(1, 31 - len(suffix))]}{suffix}"
        index += 1
    existing.add(candidate)
    return candidate


def _humanize_field_name(field_name: str) -> str:
    return field_name.replace("_", " ").strip()


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    try:
        if pd.isna(value):
            return True
    except TypeError:
        return False
    return False


def _column_has_data(series: pd.Series) -> bool:
    return any(not _is_empty_value(value) for value in series.tolist())


def _build_contract_maps(contract: ContractSpec | None) -> tuple[dict[str, list[str]], dict[str, dict[str, str]]]:
    if contract is None:
        return {}, {}

    ordered_columns: dict[str, list[str]] = {}
    labels_by_record: dict[str, dict[str, str]] = {}
    for record in contract.record_types:
        ordered_columns[record.name] = [field.name for field in record.fields]
        labels_by_record[record.name] = {
            field.name: (field.description or _humanize_field_name(field.name))
            for field in record.fields
        }
    return ordered_columns, labels_by_record


def _select_record_columns(
    group_df: pd.DataFrame,
    record_type: str,
    ordered_columns: dict[str, list[str]],
) -> pd.DataFrame:
    preferred = ordered_columns.get(record_type, [])
    selected: list[str] = []

    if "line_number" in group_df.columns:
        selected.append("line_number")

    selected.extend([field for field in preferred if field in group_df.columns])

    for column in group_df.columns:
        if column == "record_type" or column in selected:
            continue
        if _column_has_data(group_df[column]):
            selected.append(column)

    if not selected:
        selected = [column for column in group_df.columns if column != "record_type"]
    return group_df[selected].copy()


def _build_dictionary_df(contract: ContractSpec | None) -> pd.DataFrame | None:
    if contract is None:
        return None

    rows: list[dict[str, Any]] = []
    for record in contract.record_types:
        for field in record.fields:
            rows.append(
                {
                    "Zone": record.name,
                    "Champ": field.name,
                    "Libelle": field.description or "",
                    "Type": field.type.value,
                    "Debut": field.start,
                    "Longueur": field.length,
                    "Decimales": field.decimals if field.decimals is not None else "",
                }
            )
    if not rows:
        return None
    return pd.DataFrame(rows)


def _dictionary_column_labels() -> dict[str, str]:
    return {
        "Zone": "Type d'enregistrement",
        "Champ": "Nom technique du champ",
        "Libelle": "Description metier issue du programme",
        "Type": "Type de donnee",
        "Debut": "Position de debut (1-based)",
        "Longueur": "Longueur fixe",
        "Decimales": "Nombre de decimales implicites",
    }


def _infer_numeric_kind(values: list[Any]) -> str | None:
    has_decimal = False
    has_value = False
    for value in values:
        if value in (None, ""):
            continue
        has_value = True
        if isinstance(value, bool):
            return None
        if isinstance(value, Decimal):
            if value != value.to_integral_value():
                has_decimal = True
            continue
        if isinstance(value, int):
            continue
        if isinstance(value, float):
            if not float(value).is_integer():
                has_decimal = True
            continue
        if isinstance(value, str):
            text = value.strip()
            if not text:
                continue
            if len(text) > 1 and text[0] == "0" and text.isdigit():
                return None
            candidate = text.replace(" ", "").replace(",", ".")
            if re.fullmatch(r"[+-]?\d+(\.\d+)?", candidate):
                if "." in candidate:
                    has_decimal = True
                continue
            return None
        return None

    if not has_value:
        return None
    return "decimal" if has_decimal else "integer"


def _set_column_widths(worksheet) -> None:
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, worksheet.max_column + 1):
        header = worksheet.cell(4, col_idx).value
        label = worksheet.cell(3, col_idx).value
        max_len = len(str(header)) if header is not None else 0
        if label is not None:
            max_len = max(max_len, min(len(str(label)), 38))

        sample_limit = min(worksheet.max_row, 3000)
        for row_idx in range(5, sample_limit + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))

        header_name = str(header or "").upper()
        if header_name in {"EAN13", "ISBN13", "NUFAC"}:
            max_len = max(max_len, 15)
        if "LIB" in header_name:
            max_len = max(max_len, 32)
        if "DATE" in header_name or header_name.startswith("DA"):
            max_len = max(max_len, 12)

        width = min(max(max_len + 2, 10), 54)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width


def _is_currency_column(header_name: str) -> bool:
    amount_tokens = ("MON", "MTT", "NET", "PUN", "PRIX", "TOTAL", "TTC", "TVA")
    return any(token in header_name for token in amount_tokens)


def _is_quantity_column(header_name: str) -> bool:
    quantity_tokens = ("QTE", "QTFAC", "QTLIV", "QUANT")
    return any(token in header_name for token in quantity_tokens)


def _is_date_column(header_name: str) -> bool:
    return "DATE" in header_name or header_name.startswith("DA") or header_name.endswith("DAT")


def _style_data_sheet(
    worksheet,
    table_index: int,
    sheet_title: str,
    record_count: int,
    generated_at: str,
    column_labels: dict[str, str] | None = None,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if worksheet.max_row < 4 or worksheet.max_column < 1:
        return

    header_row = 4
    data_start_row = 5
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = f"A{data_start_row}"
    worksheet.row_dimensions[1].height = 28
    worksheet.row_dimensions[2].height = 20
    worksheet.row_dimensions[3].height = 32 if column_labels else 6
    worksheet.row_dimensions[4].height = 24

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    worksheet["A1"] = sheet_title
    worksheet["A2"] = f"Genere le: {generated_at} | Enregistrements: {record_count:,}"
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=13, name="Calibri")
    worksheet["A2"].font = Font(color="0B1F33", size=10, name="Calibri")
    worksheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor="0B1F33")
    worksheet["A2"].fill = PatternFill(fill_type="solid", fgColor="EAF2FB")

    header_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    label_fill = PatternFill(fill_type="solid", fgColor="EDF4FB")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    label_font = Font(color="0B1F33", size=9, italic=True)
    stripe_fill = PatternFill(fill_type="solid", fgColor="F6FAFF")

    for col_idx in range(1, max_col + 1):
        header_name = str(worksheet.cell(header_row, col_idx).value or "")
        label_text = ""
        if column_labels:
            label_text = column_labels.get(header_name, "")
        label_cell = worksheet.cell(3, col_idx)
        label_cell.value = label_text
        label_cell.font = label_font
        label_cell.fill = label_fill
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        cell = worksheet.cell(header_row, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx in range(data_start_row, max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, max_col + 1):
                worksheet.cell(row_idx, col_idx).fill = stripe_fill

    for col_idx in range(1, max_col + 1):
        sample_values = [worksheet.cell(r, col_idx).value for r in range(data_start_row, min(max_row, 1200) + 1)]
        numeric_kind = _infer_numeric_kind(sample_values)
        header_name = str(worksheet.cell(header_row, col_idx).value or "").upper()

        alignment = Alignment(horizontal="left", vertical="center")
        number_format = None
        if _is_quantity_column(header_name) or numeric_kind == "integer":
            alignment = Alignment(horizontal="right", vertical="center")
            number_format = "#,##0"
        elif _is_currency_column(header_name) or numeric_kind == "decimal":
            alignment = Alignment(horizontal="right", vertical="center")
            number_format = "#,##0.00"

        if _is_date_column(header_name):
            alignment = Alignment(horizontal="center", vertical="center")

        for row_idx in range(data_start_row, max_row + 1):
            cell = worksheet.cell(row_idx, col_idx)
            cell.alignment = alignment
            if number_format and isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = number_format

    _set_column_widths(worksheet)

    if max_row >= data_start_row:
        ref = f"A{header_row}:{get_column_letter(max_col)}{max_row}"
        worksheet.auto_filter.ref = ref


def _build_summary_sheet(
    workbook,
    all_df: pd.DataFrame,
    generated_at: str,
    contract: ContractSpec | None = None,
) -> None:
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = workbook.create_sheet(_SHEET_SUMMARY, 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws["A1"] = "IDP470 - Synthese"
    ws["A2"] = f"Genere le: {generated_at}"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A2"].font = Font(name="Calibri", size=10, color="0B1F33")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="0B1F33")
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor="EAF2FB")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 28

    total_records = int(len(all_df))

    invoice_count = 0
    if "NUFAC" in all_df.columns:
        if "record_type" in all_df.columns:
            ent = all_df[all_df["record_type"] == "ENT"]
            subset = ent if not ent.empty else all_df
        else:
            subset = all_df
        invoice_values = {
            str(v).strip()
            for v in subset["NUFAC"].tolist()
            if v is not None and str(v).strip()
        }
        invoice_count = len(invoice_values)

    metric_fill = PatternFill(fill_type="solid", fgColor="EAF2FB")
    metric_title_font = Font(name="Calibri", size=10, color="0B1F33", bold=True)
    metric_value_font = Font(name="Calibri", size=18, color="0B1F33", bold=True)

    cards = [
        ("A4:C4", "Total d'enregistrement", "A5:C6", total_records),
        ("D4:F4", "Factures", "D5:F6", invoice_count),
    ]
    for title_range, title_text, value_range, value in cards:
        ws.merge_cells(title_range)
        ws.merge_cells(value_range)
        title_cell = ws[title_range.split(":")[0]]
        value_cell = ws[value_range.split(":")[0]]
        title_cell.value = title_text
        value_cell.value = int(value)
        title_cell.font = metric_title_font
        value_cell.font = metric_value_font
        title_cell.fill = metric_fill
        value_cell.fill = metric_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.alignment = Alignment(horizontal="center", vertical="center")

    start_row = 9
    ws[f"A{start_row}"] = "Type d'enregistrement"
    ws[f"B{start_row}"] = "Nombre"
    ws[f"C{start_row}"] = "Part"
    ws[f"A{start_row}"].font = Font(bold=True, color="FFFFFF")
    ws[f"B{start_row}"].font = Font(bold=True, color="FFFFFF")
    ws[f"C{start_row}"].font = Font(bold=True, color="FFFFFF")
    ws[f"A{start_row}"].fill = PatternFill(fill_type="solid", fgColor="0F766E")
    ws[f"B{start_row}"].fill = PatternFill(fill_type="solid", fgColor="0F766E")
    ws[f"C{start_row}"].fill = PatternFill(fill_type="solid", fgColor="0F766E")
    ws[f"A{start_row}"].alignment = Alignment(horizontal="center")
    ws[f"B{start_row}"].alignment = Alignment(horizontal="center")
    ws[f"C{start_row}"].alignment = Alignment(horizontal="center")

    counts = Counter(all_df["record_type"].tolist()) if "record_type" in all_df.columns else Counter()
    current = start_row + 1
    for record_type in _sort_record_types(list(counts.keys()), contract):
        count = counts.get(record_type, 0)
        ws[f"A{current}"] = record_type
        ws[f"B{current}"] = int(count)
        ws[f"C{current}"] = (count / total_records) if total_records else 0
        ws[f"B{current}"].number_format = "#,##0"
        ws[f"C{current}"].number_format = "0.0%"
        ws[f"B{current}"].alignment = Alignment(horizontal="right")
        ws[f"C{current}"].alignment = Alignment(horizontal="right")
        current += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    logo_path = Path.cwd() / "assets" / "logo_hachette_livre.png"
    if not logo_path.exists():
        logo_path = Path(__file__).resolve().parent.parent / "assets" / "logo_hachette_livre.png"
    if logo_path.exists():
        max_width = 150
        max_height = 36
        try:
            from PIL import Image as PILImage

            with PILImage.open(logo_path) as pil_logo:
                scale = min(
                    max_width / max(pil_logo.width, 1),
                    max_height / max(pil_logo.height, 1),
                    1.0,
                )
                if scale < 1.0:
                    new_width = max(1, int(pil_logo.width * scale))
                    new_height = max(1, int(pil_logo.height * scale))
                    resample = getattr(PILImage, "Resampling", PILImage).LANCZOS
                    resized_logo = pil_logo.resize((new_width, new_height), resample)
                else:
                    resized_logo = pil_logo.copy()

                image_buffer = BytesIO()
                resized_logo.save(image_buffer, format="PNG")
                image_buffer.seek(0)
                logo = XLImage(image_buffer)
                logo.width = resized_logo.width
                logo.height = resized_logo.height
        except Exception:  # noqa: BLE001
            logo = XLImage(str(logo_path))
            scale = min(max_width / max(logo.width, 1), max_height / max(logo.height, 1), 1.0)
            logo.width = int(logo.width * scale)
            logo.height = int(logo.height * scale)
        ws.add_image(logo, "E1")


def export_to_excel(
    records: list[dict[str, Any]],
    output_path: Path,
    contract: ContractSpec | None = None,
) -> None:
    if not records:
        raise ValueError("Aucun enregistrement a exporter vers Excel.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    all_df = pd.DataFrame(records)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ordered_columns, labels_by_record = _build_contract_maps(contract)
    dictionary_df = _build_dictionary_df(contract)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        all_df.to_excel(writer, index=False, sheet_name=_SHEET_ALL, startrow=3)
        all_labels: dict[str, str] = {}
        if "record_type" in all_df.columns:
            all_labels["record_type"] = "Type d'enregistrement"
        if "line_number" in all_df.columns:
            all_labels["line_number"] = "Numero de ligne source"

        sheet_specs: list[tuple[str, pd.DataFrame, str, dict[str, str] | None]] = [
            (_SHEET_ALL, all_df, "Tous les enregistrements", all_labels or None)
        ]
        existing_sheet_names = {_SHEET_ALL, _SHEET_SUMMARY, _SHEET_DICTIONARY}

        if "record_type" in all_df.columns:
            grouped_map = {
                str(record_type).strip().upper(): group_df
                for record_type, group_df in all_df.groupby("record_type", sort=False)
            }
            for record_type_name in _sort_record_types(list(grouped_map.keys()), contract):
                group_df = grouped_map[record_type_name]
                filtered_group_df = _select_record_columns(group_df, record_type_name, ordered_columns)
                base_name = record_type_name
                safe_sheet_name = _safe_sheet_name(base_name, existing_sheet_names)
                filtered_group_df.to_excel(writer, index=False, sheet_name=safe_sheet_name, startrow=3)

                zone_labels: dict[str, str] = {}
                if "line_number" in filtered_group_df.columns:
                    zone_labels["line_number"] = "Numero de ligne source"
                zone_labels.update(labels_by_record.get(record_type_name, {}))
                for column in filtered_group_df.columns:
                    zone_labels.setdefault(column, _humanize_field_name(column))

                sheet_specs.append(
                    (
                        safe_sheet_name,
                        filtered_group_df,
                        f"Zone {record_type_name}",
                        zone_labels,
                    )
                )

        if dictionary_df is not None:
            dictionary_df.to_excel(writer, index=False, sheet_name=_SHEET_DICTIONARY, startrow=3)
            sheet_specs.append(
                (
                    _SHEET_DICTIONARY,
                    dictionary_df,
                    "Dictionnaire des donnees",
                    _dictionary_column_labels(),
                )
            )

        _build_summary_sheet(
            writer.book,
            all_df,
            generated_at=generated_at,
            contract=contract,
        )

        table_index = 1
        for sheet_name, df_sheet, title, labels in sheet_specs:
            worksheet = writer.book[sheet_name]
            _style_data_sheet(
                worksheet,
                table_index=table_index,
                sheet_title=title,
                record_count=len(df_sheet),
                generated_at=generated_at,
                column_labels=labels,
            )
            table_index += 1

    LOGGER.info("Excel exported to %s", output_path)


def _first_non_empty(record: dict[str, Any], keys: list[str]) -> Any:
    for key in keys:
        if key in record and record[key] not in ("", None):
            return record[key]
    return ""


def _signed_value(sign: Any, value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal(0)
    decimal_value = Decimal(str(value))
    return -decimal_value if str(sign).strip() == "-" else decimal_value


def _fmt_amount(value: Decimal) -> str:
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def _resolve_logo_path(logo_path: Path | None) -> Path | None:
    if logo_path and logo_path.exists():
        return logo_path

    candidates = [
        Path.cwd() / "assets" / "logo_hachette_livre.png",
        Path(__file__).resolve().parent.parent / "assets" / "logo_hachette_livre.png",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def export_first_invoice_pdf(
    records: list[dict[str, Any]],
    output_path: Path,
    logo_path: Path | None = None,
) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ModuleNotFoundError as error:
        raise RuntimeError("PDF export requires reportlab. Install with: pip install reportlab") from error

    ent_records = [record for record in records if record.get("record_type") == "ENT"]
    if not ent_records:
        raise ValueError("Aucun enregistrement ENT trouve pour construire le PDF de facture.")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(output_path), pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "invoice_title",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=16,
        textColor=colors.HexColor("#0b1f33"),
    )
    meta_style = ParagraphStyle(
        "invoice_meta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=12,
        textColor=colors.HexColor("#334155"),
    )
    story: list[Any] = []

    invoice_entries: list[tuple[str, dict[str, Any]]] = []
    seen_keys: set[str] = set()
    for index, ent in enumerate(ent_records, start=1):
        invoice_key = str(ent.get("NUFAC", "")).strip() or f"SANS_NUM_{index}"
        if invoice_key in seen_keys:
            continue
        seen_keys.add(invoice_key)
        invoice_entries.append((invoice_key, ent))

    resolved_logo = _resolve_logo_path(logo_path)
    for invoice_index, (invoice_key, ent) in enumerate(invoice_entries, start=1):
        if invoice_index > 1:
            story.append(PageBreak())

        invoice_date = str(ent.get("DAFAC", "")).strip()
        adr = next(
            (
                record
                for record in records
                if record.get("record_type") == "ADR" and str(record.get("NUFAC", "")).strip() == invoice_key
            ),
            None,
        )
        if adr is None:
            adr = next((record for record in records if record.get("record_type") == "ADR"), {})

        lig_rows = [
            record
            for record in records
            if record.get("record_type") == "LIG" and str(record.get("NUFAC", "")).strip() == invoice_key
        ]
        if not lig_rows:
            lig_rows = [record for record in records if record.get("record_type") == "LIG"][:50]

        total_ht = _signed_value(ent.get("SMONHT"), ent.get("MONHT"))
        total_tva = _signed_value(ent.get("SMTTVA"), ent.get("MTTVA"))
        total_ttc = _signed_value(ent.get("SMTTTC"), ent.get("MTTTC"))

        if resolved_logo:
            image_reader = ImageReader(str(resolved_logo))
            src_w, src_h = image_reader.getSize()
            target_w = 47 * mm
            target_h = target_w * (src_h / src_w)
            if target_h > 18 * mm:
                ratio = (18 * mm) / target_h
                target_w *= ratio
                target_h *= ratio
            header_logo: Any = Image(str(resolved_logo), width=target_w, height=target_h)
        else:
            header_logo = Paragraph("<b>Hachette Livre</b>", styles["Heading3"])

        header_info = Table(
            [
                [Paragraph("FACTURE", title_style)],
                [Paragraph(f"Numero facture: <b>{invoice_key}</b>", meta_style)],
                [Paragraph(f"Date: <b>{invoice_date}</b>", meta_style)],
            ],
            colWidths=[84 * mm],
        )
        header_info.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )

        header_table = Table([[header_logo, header_info]], colWidths=[90 * mm, 90 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#c6d3e1")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f7fbff")),
                ]
            )
        )
        story.append(header_table)
        story.append(Spacer(1, 6 * mm))

        client_name = _first_non_empty(adr, ["CLLIV_RASOC", "CLLIV_NOCLI"])
        client_address = [
            _first_non_empty(adr, ["CLLIV_ADCLI"]),
            _first_non_empty(adr, ["CLLIV_LORES"]),
            _first_non_empty(adr, ["CLLIV_LOBDI"]),
            " ".join(
                str(value).strip()
                for value in [_first_non_empty(adr, ["CLLIV_CPCLI"]), _first_non_empty(adr, ["CLLIV_CPAYS"])]
                if str(value).strip()
            ),
        ]
        lines = [str(client_name).strip(), *[str(line).strip() for line in client_address if str(line).strip()]]
        client_block = "<br/>".join(lines)
        story.append(
            Paragraph(
                f"<b>Client</b><br/>{client_block}",
                ParagraphStyle("client_block", parent=styles["Normal"], fontName="Helvetica", fontSize=9.5, leading=12),
            )
        )
        story.append(Spacer(1, 6 * mm))

        table_rows = [["Ligne", "EAN13", "Designation", "Qte", "PU HT", "Net HT"]]
        for row in lig_rows:
            description = _first_non_empty(row, ["CT_LIBTI", "LIBTI"])
            quantity = _first_non_empty(row, ["CT_QTFAC", "QTFAC"])
            unit_price = _first_non_empty(row, ["CT_PUNHT", "PUNHT"])
            net_ht = _signed_value(
                _first_non_empty(row, ["CT_SNETHT", "SNETHT"]),
                _first_non_empty(row, ["CT_NETHT", "NETHT"]),
            )
            table_rows.append(
                [
                    _first_non_empty(row, ["NULIG"]),
                    _first_non_empty(row, ["EAN13"]),
                    str(description)[:60],
                    quantity,
                    unit_price,
                    _fmt_amount(net_ht),
                ]
            )

        lines_table = Table(
            table_rows,
            colWidths=[18 * mm, 33 * mm, 70 * mm, 16 * mm, 24 * mm, 24 * mm],
            repeatRows=1,
        )
        lines_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eaf2fb")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b7c8d8")),
                    ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(lines_table)
        story.append(Spacer(1, 6 * mm))

        totals_rows = [
            ["Total HT", _fmt_amount(total_ht)],
            ["TVA", _fmt_amount(total_tva)],
            ["Total TTC", _fmt_amount(total_ttc)],
        ]
        totals_table = Table(totals_rows, colWidths=[35 * mm, 35 * mm], hAlign="RIGHT")
        totals_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#b7c8d8")),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#eaf2fb")),
                ]
            )
        )
        story.append(totals_table)

    doc.build(story)
    LOGGER.info("PDF exported to %s", output_path)
